"""Shared plumbing for the no-AI office scripts.

EVERY SCRIPT IN THIS PACKAGE IS DETERMINISTIC. No model, no network, no vendor.
They read a CSV or Excel export that the practice can already produce today,
apply arithmetic and date logic, and write a dated artifact. `make lint` asserts
there is no model import anywhere under `officeops/`.

Three design decisions carry most of the value:

  * **Column names are configuration, not code.** Every EHR exports the same
    facts under different headers -- `Appt Date`, `APPOINTMENT_DATE`,
    `appt_dt`. Hard-coding one vendor's spelling means every script has to be
    edited by an engineer the first time the practice changes systems, which is
    exactly the moment nobody has an engineer. `officeops/mappings/*.yaml` holds
    the aliases and `--mapping` overrides them.

  * **Read-only unless asked.** Nothing writes to the practice's files. Output
    goes to an output directory, dated, and `--write` is required to produce it
    at all. A script that quietly rewrites the export it was given cannot be
    run twice.

  * **Refusals are loud.** A missing column, an unparseable date, a row with no
    patient identifier -- these are reported as exceptions with row numbers, not
    dropped. A report that silently skipped 12% of its input is worse than no
    report, because somebody will act on the 88%.

ON PHI: these scripts run on the practice's own machine against the practice's
own exports and make no network calls of any kind. That is the entire privacy
architecture, and it is why this layer needs no BAA, no vendor review and no
model. Section 3.1's rules are satisfied by construction rather than by policy.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import json
import os
import re
import sys
from dataclasses import dataclass, field
from typing import Any, Callable, Iterable, Iterator, Mapping, Sequence

__all__ = [
    "OfficeOpsError",
    "MissingColumn",
    "EmptyExport",
    "TruncatedFile",
    "BadValue",
    "Row",
    "Table",
    "load_table",
    "load_mapping",
    "parse_date",
    "parse_datetime",
    "parse_float",
    "parse_bool",
    "days_between",
    "Report",
    "write_csv",
    "write_text",
    "add_common_arguments",
    "resolve_output",
    "today",
]

DEFAULT_MAPPING_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mappings")


class OfficeOpsError(RuntimeError):
    """Base for every refusal in this package."""


class MissingColumn(OfficeOpsError):
    """A required field is not present under any known alias."""


class BadValue(OfficeOpsError):
    """A cell could not be interpreted. Carries the row number."""


class EmptyExport(OfficeOpsError):
    """Headers but no rows. Not a mapping problem; do not send anyone there."""


class TruncatedFile(OfficeOpsError):
    """Fewer rows parsed than the file physically contains."""


# -- dates -------------------------------------------------------------------

#: Formats seen in real EHR exports, most specific first. US-format first
#: because that is what every product sold into a US practice emits; ISO is
#: tried before it only when the string is unambiguously ISO-shaped.
_DATE_FORMATS = (
    "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%b-%y", "%b %d %Y",
    "%B %d, %Y", "%Y/%m/%d", "%m-%d-%Y",
)
_DATETIME_FORMATS = tuple(
    f"{d} {t}" for d in _DATE_FORMATS for t in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p")
) + _DATE_FORMATS

_ISO_LIKE = re.compile(r"^\d{4}-\d{2}-\d{2}")


def parse_date(value: Any, *, field_name: str = "date", row: int | None = None) -> dt.date:
    """Parse a date, or raise naming the row.

    WHY IT RAISES: an unparseable appointment date in a confirmation-call
    export is not a row to skip. It is a patient who will not be called.
    """
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value or "").strip()
    if not text:
        raise BadValue(_where(f"{field_name} is empty", row))
    if _ISO_LIKE.match(text):
        try:
            return dt.date.fromisoformat(text[:10])
        except ValueError:
            pass
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise BadValue(_where(f"{field_name}={text!r} is not a date this tool recognises", row))


def parse_datetime(
    value: Any, *, field_name: str = "timestamp", row: int | None = None
) -> dt.datetime:
    if isinstance(value, dt.datetime):
        return value
    text = str(value or "").strip()
    if not text:
        raise BadValue(_where(f"{field_name} is empty", row))
    if _ISO_LIKE.match(text):
        try:
            return dt.datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            pass
    for fmt in _DATETIME_FORMATS:
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise BadValue(_where(f"{field_name}={text!r} is not a timestamp", row))


def parse_float(
    value: Any, *, field_name: str = "value", row: int | None = None,
    allow_blank: bool = False,
) -> float | None:
    text = str(value if value is not None else "").strip()
    # Strip the decorations spreadsheets add: currency, thousands separators,
    # degree signs, stray unit suffixes.
    cleaned = re.sub(r"[^\d.\-+eE]", "", text.replace(",", ""))
    if not cleaned:
        if allow_blank:
            return None
        raise BadValue(_where(f"{field_name} is empty", row))
    try:
        return float(cleaned)
    except ValueError as exc:
        raise BadValue(
            _where(f"{field_name}={text!r} is not a number", row)
        ) from exc


_TRUE = {"y", "yes", "true", "t", "1", "confirmed", "complete", "completed", "done"}
_FALSE = {"n", "no", "false", "f", "0", "", "pending", "none", "not confirmed"}


def parse_bool(value: Any, *, field_name: str = "flag", row: int | None = None) -> bool:
    text = str(value if value is not None else "").strip().lower()
    if text in _TRUE:
        return True
    if text in _FALSE:
        return False
    raise BadValue(_where(f"{field_name}={value!r} is not yes/no", row))


def days_between(earlier: dt.date, later: dt.date) -> int:
    return (later - earlier).days


def today() -> dt.date:
    return dt.date.today()


_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def _normal(text: str) -> str:
    """Lowercase, strip, and collapse every non-alphanumeric run to nothing."""
    return _NON_ALNUM.sub("", str(text).strip().lower())


def _where(message: str, row: int | None) -> str:
    return message if row is None else f"row {row}: {message}"


# -- tables ------------------------------------------------------------------


@dataclass
class Row:
    """One export row, addressed by canonical field name rather than header."""

    number: int
    raw: Mapping[str, str]
    _aliases: Mapping[str, Sequence[str]]

    def has(self, field_name: str) -> bool:
        return self._column(field_name) is not None

    def _column(self, field_name: str) -> str | None:
        """Find the header for a canonical field, comparing on a normal form.

        Normalising rather than comparing raw strings is what keeps the mapping
        files short: `Appt Date/Time`, `appt_date_time` and `Appt Date Time` are
        one alias, not three, and a placeholder written `{{last_well_visit}}`
        matches a column headed `Last Well Visit` without anyone having to
        notice that they differ.
        """
        wanted = [_normal(a) for a in self._aliases.get(field_name, ())]
        wanted.append(_normal(field_name))
        for alias in wanted:
            for header in self.raw:
                if _normal(header) == alias:
                    return header
        return None

    def get(self, field_name: str, default: Any = None) -> Any:
        column = self._column(field_name)
        if column is None:
            return default
        return self.raw[column]

    def text(self, field_name: str, *, required: bool = True, default: str = "") -> str:
        column = self._column(field_name)
        if column is None:
            if required:
                raise MissingColumn(
                    f"no column for {field_name!r}; looked for "
                    f"{list(self._aliases.get(field_name, (field_name,)))} among "
                    f"{sorted(self.raw)}"
                )
            return default
        value = str(self.raw[column] or "").strip()
        if not value and required:
            raise BadValue(_where(f"{field_name} is empty", self.number))
        return value or default

    def date(self, field_name: str, *, required: bool = True) -> dt.date | None:
        value = self.text(field_name, required=required)
        if not value:
            return None
        return parse_date(value, field_name=field_name, row=self.number)

    def datetime(self, field_name: str, *, required: bool = True) -> dt.datetime | None:
        value = self.text(field_name, required=required)
        if not value:
            return None
        return parse_datetime(value, field_name=field_name, row=self.number)

    def number_(self, field_name: str, *, required: bool = True) -> float | None:
        value = self.text(field_name, required=required)
        if not value:
            return None
        return parse_float(value, field_name=field_name, row=self.number)

    def flag(self, field_name: str, *, default: bool = False) -> bool:
        value = self.text(field_name, required=False)
        if not value:
            return default
        return parse_bool(value, field_name=field_name, row=self.number)


@dataclass
class Table:
    """A loaded export plus the aliases used to address it."""

    rows: list[Row]
    headers: list[str]
    source: str
    aliases: Mapping[str, Sequence[str]] = field(default_factory=dict)

    def __iter__(self) -> Iterator[Row]:
        return iter(self.rows)

    def __len__(self) -> int:
        return len(self.rows)

    def require_rows(self) -> None:
        """An export with headers and no data is a different problem.

        Reporting it as a missing column sent an office manager into the mapping
        file for an afternoon over a holiday Monday when the schedule was empty.
        """
        if not self.rows:
            raise EmptyExport(
                f"{self.source} has a valid header row and no data. Either the "
                "report produced nothing for this period, or the export job did "
                "not run."
            )

    def has_column(self, field_name: str) -> bool:
        """Whether a column exists, independent of whether any row has a value.

        The distinction matters more than it looks. `text(..., required=False)`
        returns the same empty string for "this export has no confirmation
        column" and "this patient has not confirmed" -- and the first produced a
        call list where every patient read NOT CONFIRMED and the confirmation
        rate read 0%, a number the phasing plan then hands to I-07 as a measured
        baseline.
        """
        wanted = {_normal(a) for a in self.aliases.get(field_name, ())}
        wanted.add(_normal(field_name))
        return any(_normal(header) in wanted for header in self.headers)

    def require(self, *field_names: str) -> None:
        """Fail before doing any work if a needed column is absent.

        Checked up front rather than on first use: discovering at row 400 that
        the export has no `result_date` column means 399 rows of output that
        have to be thrown away, and somebody will keep the first page.
        """
        if not self.headers:
            raise MissingColumn(f"{self.source}: the file has no header row")
        missing = [name for name in field_names if not self.has_column(name)]
        if missing:
            looked_for = {
                name: list(self.aliases.get(name, [])) or [name] for name in missing
            }
            raise MissingColumn(
                f"{self.source}: no column for {missing}. Looked for "
                f"{looked_for}. Headers present: {self.headers}. Add an alias to "
                "the mapping file (see officeops/mappings/) rather than renaming "
                "the export -- the export is regenerated every day by somebody "
                "who will not remember to rename it."
            )


def _read_alias_file(path: str) -> dict[str, list[str]]:
    import yaml

    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = yaml.safe_load(handle) or {}
    except yaml.YAMLError as exc:
        raise OfficeOpsError(
            f"{os.path.basename(path)} is not valid YAML: {exc}"
        ) from exc
    columns = data.get("columns") if isinstance(data, dict) else None
    if columns is None:
        raise OfficeOpsError(
            f"{os.path.basename(path)} has no top-level 'columns:' mapping. "
            "Expected:\n\ncolumns:\n  patient_name: [Patient, PT_NAME]"
        )
    if not isinstance(columns, dict):
        raise OfficeOpsError(
            f"{os.path.basename(path)}: 'columns' must be a mapping of "
            f"field -> list of headers, not {type(columns).__name__}"
        )
    return {
        str(key): [str(v) for v in (value if isinstance(value, list) else [value])]
        for key, value in columns.items()
    }


def load_mapping(
    name: str, *, directory: str = DEFAULT_MAPPING_DIR, base: str | None = None
) -> dict[str, list[str]]:
    """Load column aliases, MERGED over the shared defaults.

    Merging rather than replacing is what makes the documented workflow work. A
    practice told to "copy the mapping and add your header" writes a four-key
    file for the four columns that differ -- and a replacing loader then dropped
    `patient_id` and `phone`, producing a confirmation call list with no phone
    numbers on it and nothing saying why.

    `_common.yaml` is loaded first for the same reason: it is the file whose own
    header tells you to edit it, so it has to actually be read.
    """
    merged: dict[str, list[str]] = {}
    for source in (os.path.join(directory, "_common.yaml"), base):
        if source and os.path.exists(source):
            merged.update(_read_alias_file(source))
    path = name if (os.path.sep in name or name.endswith((".yaml", ".yml"))) \
        else os.path.join(directory, f"{name}.yaml")
    if os.path.exists(path):
        merged.update(_read_alias_file(path))
    elif os.path.sep in name or name.endswith((".yaml", ".yml")):
        raise OfficeOpsError(f"mapping file not found: {name}")
    return merged


def load_table(
    path: str | os.PathLike[str],
    *,
    aliases: Mapping[str, Sequence[str]] | None = None,
    encoding: str = "utf-8-sig",
) -> Table:
    """Read a CSV (or a single-sheet XLSX) into canonical rows.

    `utf-8-sig` by default because Excel writes a BOM and a BOM on the first
    header turns `Patient ID` into `\\ufeffPatient ID`, which then matches no
    alias and produces a "no column for patient_id" refusal on a file that
    plainly has one.
    """
    path = str(path)
    aliases = dict(aliases or {})
    if path.lower().endswith((".xlsx", ".xlsm")):
        headers, records = _read_xlsx(path)
    else:
        with open(path, "r", encoding=encoding, newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            records = [dict(r) for r in reader]
        _assert_not_truncated(path, encoding, len(records), len(headers))
    rows = [
        Row(number=index + 2, raw=record, _aliases=aliases)
        for index, record in enumerate(records)
    ]
    return Table(rows=rows, headers=headers, source=os.path.basename(path), aliases=aliases)


def _assert_not_truncated(path: str, encoding: str, parsed: int, columns: int) -> None:
    """Refuse a file whose parsed row count is far below its physical line count.

    One unbalanced quote makes `csv` swallow everything after it. A twenty-row
    schedule export became six rows, the report said "5 of 5 appointments are
    not confirmed", and fifteen families were never called -- with the tool
    reporting one unreadable row. This is the exact failure the loud-refusal
    rule exists to prevent, and only a physical line count can see it.

    The tolerance is generous because a quoted field may legitimately contain
    newlines; what it catches is an order-of-magnitude discrepancy.
    """
    with open(path, "r", encoding=encoding, newline="") as handle:
        physical = sum(1 for _ in handle)
    data_lines = max(physical - 1, 0)
    if data_lines and parsed < data_lines * 0.75:
        raise TruncatedFile(
            f"{os.path.basename(path)}: parsed {parsed} row(s) from a file with "
            f"{data_lines} data line(s). An unterminated quote makes the CSV "
            "reader swallow everything after it. Fix the export before trusting "
            "any report built from it."
        )


def _read_xlsx(path: str) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise OfficeOpsError(
            f"{path} is an Excel file and openpyxl is not installed. Either "
            "`pip install openpyxl` or re-export as CSV -- every EHR can."
        ) from exc
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(h or "") for h in next(iterator, ())]
    records = [
        {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(values) if i < len(headers)}
        for values in iterator
        if any(v is not None and str(v).strip() for v in values)
    ]
    return headers, records


# -- output ------------------------------------------------------------------


@dataclass
class Report:
    """What a task produces: a headline, counts, findings, and refusals.

    `problems` is separate from `findings` and is printed even when empty,
    because "0 rows could not be read" is information and a silent absence is
    not.
    """

    task: str
    generated: dt.datetime
    headline: str = ""
    counts: dict[str, Any] = field(default_factory=dict)
    findings: list[dict[str, Any]] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)
    columns: list[str] = field(default_factory=list)
    #: The command and the inputs that produced this. A CSV reopened in three
    #: months with no as-of date, no threshold and no source filename cannot be
    #: dated or reproduced, and a retention policy that keeps it is keeping an
    #: undated list of patients.
    parameters: dict[str, Any] = field(default_factory=dict)
    sources: list[str] = field(default_factory=list)
    #: Rendered documents, written only to their own file. Never into the report
    #: body or the JSON -- that duplicated PHI onto disk in an artifact the
    #: retention policy did not know about.
    letters: list[str] = field(default_factory=list)

    def add(self, **fields: Any) -> None:
        self.findings.append(fields)
        for key in fields:
            if key not in self.columns:
                self.columns.append(key)

    def problem(self, message: str) -> None:
        self.problems.append(message)

    def as_dict(self) -> dict[str, Any]:
        return {
            "task": self.task,
            "generated": self.generated.isoformat(timespec="seconds"),
            "headline": self.headline,
            "parameters": dict(self.parameters),
            "sources": list(self.sources),
            "counts": {k: v for k, v in self.counts.items() if not k.startswith("_")},
            "findings": list(self.findings),
            "problems": list(self.problems),
        }

    def render(self, *, width: int = 78, limit: int = 40) -> str:
        rule = "=" * width
        out = [rule, f"{self.task}  ({self.generated:%Y-%m-%d %H:%M})", rule]
        if self.sources:
            out.append(f"  source: {', '.join(self.sources)}")
        if self.parameters:
            out.append(
                "  parameters: "
                + ", ".join(f"{k}={v}" for k, v in sorted(self.parameters.items()))
            )
        if self.headline:
            out += ["", self.headline]
        if self.counts:
            out.append("")
            for key, value in self.counts.items():
                if key.startswith("_"):
                    continue
                out.append(f"  {key.replace('_', ' '):<38} {value}")
        if self.findings:
            out += ["", f"{len(self.findings)} item(s):", ""]
            widths = {
                c: max(len(c), *(len(str(f.get(c, ""))) for f in self.findings))
                for c in self.columns
            }
            out.append("  " + "  ".join(c.replace("_", " ").upper().ljust(widths[c]) for c in self.columns))
            out.append("  " + "  ".join("-" * widths[c] for c in self.columns))
            for finding in self.findings[:limit]:
                out.append(
                    "  " + "  ".join(str(finding.get(c, "")).ljust(widths[c]) for c in self.columns)
                )
            if len(self.findings) > limit:
                out.append(f"  ... and {len(self.findings) - limit} more (see the CSV)")
        else:
            out += ["", "Nothing to report."]
        out += ["", f"{len(self.problems)} row(s) could not be read."]
        for problem in self.problems[:20]:
            out.append(f"  ! {problem}")
        if len(self.problems) > 20:
            out.append(f"  ! ... and {len(self.problems) - 20} more")
        return "\n".join(out)


def write_csv(report: Report, path: str | os.PathLike[str]) -> str:
    """Findings, prefixed by the run that produced them.

    Three provenance columns on every row rather than a header block, because a
    CSV gets opened in Excel, sorted, filtered and pasted into an email -- and a
    header block does not survive any of that while a column does.
    """
    os.makedirs(os.path.dirname(os.path.abspath(str(path))) or ".", exist_ok=True)
    provenance = ["run_task", "run_as_of", "run_parameters"]
    stamp = {
        "run_task": report.task,
        "run_as_of": report.parameters.get("as_of")
        or report.generated.date().isoformat(),
        "run_parameters": "; ".join(
            f"{k}={v}" for k, v in sorted(report.parameters.items())
        ),
    }
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=provenance + (report.columns or ["note"])
        )
        writer.writeheader()
        for finding in report.findings:
            writer.writerow({**stamp, **finding})
    return str(path)


def write_text(text: str, path: str | os.PathLike[str]) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(str(path))) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as handle:
        handle.write(text)
    return str(path)


def add_common_arguments(parser: argparse.ArgumentParser) -> argparse.ArgumentParser:
    parser.add_argument("--out", default="out", help="output directory (default: ./out)")
    parser.add_argument(
        "--write", action="store_true",
        help="actually write the output files (default: print only)",
    )
    parser.add_argument("--json", action="store_true", help="print the report as JSON")
    parser.add_argument(
        "--mapping", default=None,
        help="path to a column-alias YAML overriding officeops/mappings/",
    )
    parser.add_argument(
        "--today", default=None,
        help="treat this ISO date as today (for testing and for re-running a past day)",
    )
    return parser


def resolve_output(args: argparse.Namespace, filename: str) -> str:
    return os.path.join(args.out, filename)
